import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill,Font,Border,Side 
import plotly.express as px

def scan_folder(folder_path):
    folder = Path(folder_path)
    valid_files = []
    report = []

    if not folder.exists():
        print(f"路径不存在：{folder_path}")
        return None
    
    for file in folder.glob("*.xls*"):
        try:
            df = pd.read_excel(file)
        except Exception as e:
            report.append({'file': file.name, 'issue': f"无法打开:{e}"})
            continue
        if df.empty:
            report.append({'file': file.name, 'issue': "文件为空"})
            continue       
        valid_files.append(str(file))   
    return {
        "valid_files": valid_files,
        "report": report
    }

def read_and_align(file_path, standard_cols, rename_map=None):
    '''
    读取函数接收一个文件地址来进行转换和过滤处理
    standar_cols 标准表头值
    rename_map 映射字典
    '''
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        print(f"无法读取文件 {file_path}: {e}")
        return None
    if df.empty:
        print(f"{file_path} 是空文件")
        return None
    if rename_map:
        df = df.rename(columns=rename_map)
    # 重命名之后再检查列名是否齐全
    missing_cols = [col for col in standard_cols if col not in df.columns]
    if missing_cols:
        print(f"{file_path} 缺少列: {missing_cols}")
        return None
    return df[standard_cols]
    
def merge_all(df_list):
    '''
    对有效文件列表做合并处理
    df_list : 有效文件列表
    ignore_index : 行号重置
    '''
    merged=pd.concat(df_list, ignore_index=True)
    return merged

def clean_data(df,col_name,subset=None,dropna_cols=None):
    '''
    对列表进行清洗,建立列表副本后进行查重-修改格式 错误值转NaN-删除空白值，最后输出处理报告
    df : 有效列表文件
    col : 格式转换行名
    subset : 可调只读列查重 默认值关闭
    dropna_cols:可调只删除某行空白值
    输出处理报告 : 原始{before}行->保留{after}行，删除{before-after}行
    '''
    df=df.copy()
    before=len(df)

    df=df.drop_duplicates(subset=subset)
    df[col_name]=pd.to_numeric(df[col_name], errors='coerce')
    df=df.dropna(subset=dropna_cols)
    df[col_name]=df[col_name].astype(int)
    
    after = len(df)
    print(f'清洗完成：原始{before}行->保留{after}行，删除{before-after}行')
    return df

def export_excel(df,output_path,header_color='4F81BD'):
    '''
    对有效excel列表文件处理,在需要的时候用openpyxl写入模式打开文件,然后美化表格,最后保存到输出文件夹里
    df : 有效列表文件名
    output_path : 输出文件夹
    header_color : 表头颜色 默认蓝色
    '''
    with pd.ExcelWriter(output_path,engine='openpyxl') as writer:
        df.to_excel(writer,index=False,sheet_name='汇总')
    wb=openpyxl.load_workbook(output_path)
    ws=wb.active

    fill=PatternFill(fill_type='solid',fgColor=header_color)
    font=Font(bold=True,color='FFFFFF')
    thin=Side(style='thin')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    ws.freeze_panes = 'A2'

    for cell in ws[1]:
        cell.fill=fill
        cell.font=font
        cell.border=border
    for row_index,row in enumerate(ws.iter_rows(min_row=2),1):
        for cell in row:
            cell.border = border
            if row_index % 2:
                cell.fill = PatternFill(fill_type='solid',fgColor='E8F4FD')

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length,len(str(cell.value)))
        ws.column_dimensions[col_letter].width=max_length+4
    wb.save(output_path)
    print(f'已导出:{output_path}')

def plot_bar(df,output_folder,filename,x,y,color=None,title=None):
    '''
    自动绘制柱状图 可选形参有颜色分类color= 表名注释title=
    '''
    output_path=Path(output_folder)/(filename+ '.html')
    fig=px.bar(df,x=x,y=y,color=color,title=title)
    fig.write_html(output_path)
    print(f'导出成功')